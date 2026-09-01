"""
C4 Agent L2 功能测试共享辅助模块 — test_helpers.py

提供:
  - create_test_csv(): 创建标准 Modbus 点表 CSV 文件
  - create_test_xlsx(): 创建标准 Modbus 点表 XLSX 文件（需 openpyxl）
  - create_csv_missing_ip(): 创建缺少 IP 字段的点表
  - create_messy_csv(): 创建混乱的点表（字段名拼写错误等）
  - create_binary_file(): 创建非解析格式的二进制文件
  - create_corrupted_xlsx(): 创建损坏的 xlsx 文件
  - retry_llm(): LLM 容忍重试装饰器（最多 3 次）
  - find_interrupt_id(): 从 SSE 流中提取 interrupt 事件 ID
  - full_access_flow(): 执行完整的接入流程（上传→解析→方案→确认→执行）
"""

import functools
import json
import os
import struct
import time
from pathlib import Path
from typing import Any, Optional, Tuple

import pytest  # type: ignore


# ──────────────────────────────────────────────
#  LLM 容忍重试装饰器
# ──────────────────────────────────────────────
# 按 README §4.1: 失败时重试最多 2 次（共 3 次尝试）


def retry_llm(max_attempts: int = 3, delay: float = 2.0):
    """
    对非确定性 LLM 输出失败进行容忍重试。

    用法:
        @retry_llm(max_attempts=3)
        def test_something(chat):
            ...
    """

    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except (AssertionError, Exception) as e:
                    last_exception = e
                    if attempt < max_attempts:
                        time.sleep(delay * attempt)  # 递增退避
                        # 重置状态：清空对话历史 + 持久化文件（config/abbr_registry），
                        # 避免重试时历史与文件残留导致 LLM 命中旧记录而困惑
                        for a in args:
                            if hasattr(a, "_history"):
                                a._history = []
                            config_dir = getattr(a, "config_dir", None)
                            if isinstance(config_dir, Path):
                                for name in (
                                    "config.json",
                                    "abbr_registry.json",
                                    "config.json.bak",
                                    "config.json.tmp",
                                ):
                                    p = config_dir / name
                                    if p.exists():
                                        try:
                                            p.unlink()
                                        except OSError:
                                            pass
                        continue
                    raise
            # unreachable 原因：最后一次循环总会 raise 或 return
            if last_exception:
                raise last_exception
            return None

        return wrapper

    return decorator


# ──────────────────────────────────────────────
#  测试数据工厂
# ──────────────────────────────────────────────

# 标准 Modbus 点表数据 — 华能阿拉善 1# 风机
STANDARD_POINTS = [
    "name,addr,uid,fun,type,swap",
    "windspeed,1000,1,3,10,2",
    "temperature,1002,1,3,10,2",
    "power,1004,1,3,10,2",
    "pressure,1006,1,3,10,2",
    "vibration,1008,1,3,10,2",
]

# 带 IP 和协议字段的完整点表
FULL_POINTS = [
    "device_name,device_ip,protocol,port,point_name,addr,uid,fun,type,swap",
    "华能阿拉善1#风机,192.168.110.1,modbus,502,windspeed,1000,1,3,10,2",
    "华能阿拉善1#风机,192.168.110.1,modbus,502,temperature,1002,1,3,10,2",
    "华能阿拉善1#风机,192.168.110.1,modbus,502,power,1004,1,3,10,2",
    "华能阿拉善1#风机,192.168.110.1,modbus,502,pressure,1006,1,3,10,2",
    "华能阿拉善1#风机,192.168.110.1,modbus,502,vibration,1008,1,3,10,2",
]

# 缺少 IP 地址字段的点表
MISSING_IP_POINTS = [
    "device_name,protocol,point_name,addr,fun,type,swap",
    "华能阿拉善1#风机,modbus,windspeed,1000,3,10,2",
    "华能阿拉善1#风机,modbus,temperature,1002,3,10,2",
]

# 混乱的点表（字段名拼写错误、数值越界等）
MESSY_POINTS = [
    "devce_naem,ip_addr,potocol,potr,pont_nam,adres,unt,funct,typ,swp",
    "hnals_1_wt,???,,99999,spd_1,0x3E8,-,ff,99,9",
    ",,modbs,,,,,,,",
    "X" * 500,
]

# CSV 格式点表（标准 Modbus 点表）
CSV_POINTS = [
    "name,addr,uid,fun,type,swap",
    "windspeed,1000,1,3,10,2",
    "temperature,1002,1,3,10,2",
]

# 协议歧义点表 — 字段不含任何协议特征列（无 uid/fun/type/swap，无 protocol 列），
# 用于 §4.3.7/4.3.8：仅凭点表字段无法唯一确定协议。
AMBIGUOUS_POINTS = [
    "name,addr,data_type",
    "windspeed,1000,uint16",
    "temperature,1002,uint16",
    "power,1004,uint16",
]

# ASFP2 数据点表 — 接收端设备 + 转发
ASFP2_POINTS = [
    "device_name,device_ip,protocol,port,point_name,addr",
    "华能阿拉善ASFP2数据源,172.16.109.11,asfp2,9999,windspeed,1000",
    "华能阿拉善ASFP2数据源,172.16.109.11,asfp2,9999,temperature,1002",
    "华能阿拉善ASFP2数据源,172.16.109.11,asfp2,9999,power,1004",
    "华能阿拉善ASFP2数据源,172.16.109.11,asfp2,9999,pressure,1006",
    "华能阿拉善ASFP2数据源,172.16.109.11,asfp2,9999,vibration,1008",
]


def create_asfp2_csv(target_dir: Path, filename: str = "asfp2_points.csv") -> Path:
    """创建 ASFP2 数据源点表 CSV。protocol=asfp2，port=9999。"""
    filepath = target_dir / filename
    filepath.write_text("\n".join(ASFP2_POINTS), encoding="utf-8")
    return filepath


def create_test_csv(target_dir: Path, filename: str = "test_points.csv") -> Path:
    """
    创建标准 Modbus 点表 CSV 文件。

    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text("\n".join(STANDARD_POINTS), encoding="utf-8")
    return filepath


def create_full_csv(
    target_dir: Path,
    filename: str = "full_points.csv",
    device_name: str = "华能阿拉善1#风机",
    device_ip: str = "192.168.110.1",
) -> Path:
    """
    创建含设备名、IP、协议的完整点表 CSV。

    device_name / device_ip 可覆盖，用于生成不同设备的点表（模拟多设备接入）。
    返回: 文件路径
    """
    rows = [
        "device_name,device_ip,protocol,port,point_name,addr,uid,fun,type,swap",
    ]
    for point_name, addr in (
        ("windspeed", 1000),
        ("temperature", 1002),
        ("power", 1004),
        ("pressure", 1006),
        ("vibration", 1008),
    ):
        rows.append(
            f"{device_name},{device_ip},modbus,502,{point_name},{addr},1,3,10,2"
        )
    filepath = target_dir / filename
    filepath.write_text("\n".join(rows), encoding="utf-8")
    return filepath


def create_csv_missing_ip(target_dir: Path, filename: str = "missing_ip.csv") -> Path:
    """
    创建缺少 IP 字段的点表 CSV。

    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text("\n".join(MISSING_IP_POINTS), encoding="utf-8")
    return filepath


def create_messy_csv(target_dir: Path, filename: str = "messy_points.csv") -> Path:
    """
    创建内容混乱的点表 CSV（字段名拼写错误、数值越界）。

    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text("\n".join(MESSY_POINTS), encoding="utf-8")
    return filepath


def create_simple_csv(target_dir: Path, filename: str = "simple_points.csv") -> Path:
    """
    创建简洁格式的点表 CSV（name,addr,uid,fun,type,swap）。

    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text("\n".join(CSV_POINTS), encoding="utf-8")
    return filepath


def create_ambiguous_csv(
    target_dir: Path, filename: str = "ambiguous_points.csv"
) -> Path:
    """
    创建协议歧义点表 CSV — 字段不含任何协议特征列（无 uid/fun/type/swap，无 protocol）。

    用于 README §4.3.7/4.3.8：仅凭点表字段无法唯一确定协议。
    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text("\n".join(AMBIGUOUS_POINTS), encoding="utf-8")
    return filepath


def create_test_xlsx(target_dir: Path, filename: str = "test_points.xlsx") -> Path:
    """
    创建标准 Modbus 点表 XLSX 文件。

    依赖 openpyxl。若未安装则 pytest.skip。
    返回: 文件路径
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        pytest.skip("openpyxl not installed — skipping xlsx test data creation")

    filepath = target_dir / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Points"

    # 写入表头
    headers = ["name", "addr", "uid", "fun", "type", "swap"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据行
    data_rows = [
        ["windspeed", 1000, 1, 3, 10, 2],
        ["temperature", 1002, 1, 3, 10, 2],
        ["power", 1004, 1, 3, 10, 2],
        ["pressure", 1006, 1, 3, 10, 2],
        ["vibration", 1008, 1, 3, 10, 2],
    ]
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(filepath))
    return filepath


def create_full_xlsx(target_dir: Path, filename: str = "full_points.xlsx") -> Path:
    """
    创建含设备名、IP、协议的完整点表 XLSX。

    依赖 openpyxl。若未安装则 pytest.skip。
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        pytest.skip("openpyxl not installed — skipping xlsx test data creation")

    filepath = target_dir / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Points"

    headers = ["device_name", "device_ip", "protocol", "port",
               "point_name", "addr", "uid", "fun", "type", "swap"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    data = [
        ["华能阿拉善1#风机", "192.168.110.1", "modbus", 502, "windspeed", 1000, 1, 3, 10, 2],
        ["华能阿拉善1#风机", "192.168.110.1", "modbus", 502, "temperature", 1002, 1, 3, 10, 2],
        ["华能阿拉善1#风机", "192.168.110.1", "modbus", 502, "power", 1004, 1, 3, 10, 2],
        ["华能阿拉善1#风机", "192.168.110.1", "modbus", 502, "pressure", 1006, 1, 3, 10, 2],
        ["华能阿拉善1#风机", "192.168.110.1", "modbus", 502, "vibration", 1008, 1, 3, 10, 2],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(filepath))
    return filepath


def create_missing_ip_xlsx(target_dir: Path, filename: str = "missing_ip.xlsx") -> Path:
    """
    创建缺少 IP 地址字段的点表 XLSX。

    依赖 openpyxl。若未安装则 pytest.skip。
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        pytest.skip("openpyxl not installed — skipping xlsx test data creation")

    filepath = target_dir / filename
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["device_name", "protocol", "point_name", "addr", "fun", "type", "swap"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    data = [
        ["华能阿拉善1#风机", "modbus", "windspeed", 1000, 3, 10, 2],
        ["华能阿拉善1#风机", "modbus", "temperature", 1002, 3, 10, 2],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(filepath))
    return filepath


def create_corrupted_xlsx(target_dir: Path, filename: str = "corrupted.xlsx") -> Path:
    """
    创建损坏的 xlsx 文件（截断 ZIP 内容）。

    先创建一个有效 xlsx，然后截断为随机长度。
    若 openpyxl 不可用，则创建随机二进制文件模拟损坏。
    """
    filepath = target_dir / filename
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="test")
        wb.save(str(filepath))
    except ImportError:
        # 无 openpyxl 时用随机二进制模拟
        filepath.write_bytes(b"PK\x03\x04" + os.urandom(200))
        return filepath

    # 截断文件 — xlsx 是 ZIP，截断会损坏 ZIP 结构
    raw = filepath.read_bytes()
    truncated = raw[: max(len(raw) // 3, 8)]
    filepath.write_bytes(truncated)
    return filepath


def create_binary_file(target_dir: Path, filename: str = "binary_data.bin") -> Path:
    """
    创建二进制文件（模拟不支持的格式）。

    返回: 文件路径
    """
    filepath = target_dir / filename
    # 写入随机二进制数据 + 非 xlsx/csv 的文件头
    content = bytearray()
    content.extend(b"\x00\x01\x02\x03")
    content.extend(os.urandom(256))
    filepath.write_bytes(bytes(content))
    return filepath


def create_text_file(target_dir: Path, filename: str = "notes.txt") -> Path:
    """
    创建不支持的文本文件格式。

    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text(
        "这是风机的备注信息\n设备名称: 华能阿拉善1#风机\n位置: 阿拉善\n",
        encoding="utf-8",
    )
    return filepath


def create_device_txt(target_dir: Path, filename: str = "device_info.txt") -> Path:
    """
    创建包含设备信息的 txt 文件（模拟工程师手写描述）。

    返回: 文件路径
    """
    filepath = target_dir / filename
    filepath.write_text(
        "设备名称: 华能阿拉善1#风机\n"
        "通信协议: Modbus TCP\n"
        "IP地址: 192.168.110.1\n"
        "端口: 502\n"
        "\n"
        "数据点列表:\n"
        "  windspeed    地址:1000  功能码:3  数据类型:uint16\n"
        "  temperature  地址:1002  功能码:3  数据类型:uint16\n"
        "  power        地址:1004  功能码:3  数据类型:uint16\n"
        "  pressure     地址:1006  功能码:3  数据类型:uint16\n"
        "  vibration    地址:1008  功能码:3  数据类型:uint16\n",
        encoding="utf-8",
    )
    return filepath


# ──────────────────────────────────────────────
#  对话流辅助
# ──────────────────────────────────────────────


def find_interrupt_id(stream: Any) -> Optional[str]:
    """
    从 SSE 流中提取 interrupt 事件的 ID。

    参数:
        stream: SSEEventStream（已消费完毕）

    返回:
        interrupt_id 字符串，若无则返回 None。
    """
    for event in stream.events:
        if event.type == "interrupt":
            try:
                data = json.loads(event.data)
                if isinstance(data, dict):
                    return data.get("id") or data.get("interrupt_id")
            except (json.JSONDecodeError, TypeError):
                pass
    return None


def get_all_text(stream: Any) -> str:
    """
    获取 SSE 流中所有文本内容。

    返回: 拼接后的完整文本。
    """
    return stream.text_content()


def run_chat(chat_helper: Any, message: str) -> Tuple[Any, str]:
    """
    便捷方法：发送一条消息并返回 (stream, text)。

    用法:
        stream, text = run_chat(chat, "你好")
    """
    with chat_helper.send(message) as stream:
        text = stream.text_content()
    return stream, text


def run_upload(chat_helper: Any, file_path: str, message: str) -> Tuple[Any, str]:
    """
    便捷方法：上传文件并返回 (stream, text)。

    用法:
        stream, text = run_upload(chat, "/path/to/points.csv", "接入此设备")
    """
    with chat_helper.send_with_file(message, file_path) as stream:
        text = stream.text_content()
    chat_helper.record_response(text)
    return stream, text


# ──────────────────────────────────────────────
#  完整接入流程辅助
# ──────────────────────────────────────────────


def _parse_csv_to_device_json(file_path: str) -> dict:
    """
    从 CSV 文件直接解析设备 JSON（不依赖 SSE event 时序）。

    支持两种格式：
    1. FULL_POINTS: device_name,device_ip,protocol,port,point_name,addr,uid,fun,type,swap
    2. STANDARD_POINTS: name,addr,uid,fun,type,swap（无设备信息列）
    """
    import csv
    import re

    SITE_ABBR = {"华能阿拉善": "hnals"}

    devices: dict[str, dict] = {}
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return {}

        has_dev_fields = any(f in (reader.fieldnames or []) for f in ["device_name", "device_ip", "protocol"])

        for row in reader:
            if has_dev_fields:
                dev_name = row.get("device_name", "") or row.get("设备名称", "")
                dev_ip = row.get("device_ip", "") or row.get("device_ip", "")
                protocol = row.get("protocol", "") or row.get("协议", "")
                port = int(row.get("port", "0") or "0")
                point_name = row.get("point_name", "") or row.get("点名称", "")
            else:
                dev_name = "设备"
                dev_ip = ""
                protocol = ""
                port = 0
                point_name = row.get("name", "") or row.get("点名称", "")

            addr = int(row.get("addr", "0") or row.get("地址", "0") or "0")
            uid = int(row.get("uid", "0") or "0") if row.get("uid") else None
            fun = int(row.get("fun", "0") or "0") if row.get("fun") else None
            typ = int(row.get("type", "0") or "0") if row.get("type") else None
            swap = int(row.get("swap", "0") or "0") if row.get("swap") else None

            if dev_name not in devices:
                seq_match = re.search(r"(\d+)#", dev_name)
                seq = int(seq_match.group(1)) if seq_match else 1
                devices[dev_name] = {
                    "name": dev_name,
                    "seq": seq,
                    "protocol": protocol,
                    "connection": {"ip": dev_ip, "port": port or 502},
                    "points": [],
                }
            pt = {"name": point_name or f"point_{len(devices[dev_name]['points'])}", "addr": addr}
            if uid is not None: pt["uid"] = uid
            if fun is not None: pt["fun"] = fun
            if typ is not None: pt["type"] = typ
            if swap is not None: pt["swap"] = swap
            devices[dev_name]["points"].append(pt)

    dev_list = list(devices.values())
    result: dict = {"devices": dev_list}
    # 从设备名提取场站信息（如 "华能阿拉善2#风机" → site 华能阿拉善, abbr hnals），
    # 使 output_plan_steps 能确定性生成 instance.id（否则 fallback 无法推断 site.abbr）。
    if dev_list:
        first_name = dev_list[0].get("name", "")
        m = re.search(r"^(.+?)\d+#", first_name)
        if m:
            site_name = m.group(1)
            abbr = SITE_ABBR.get(site_name)
            if abbr:
                result["site"] = {"name": site_name, "abbr": abbr}
    return result


def full_access_flow(
    chat: Any,
    agent: Any,
    file_path: str,
    upload_msg: str,
    plan_msg: str = "生成接入方案，并转发到中心侧",
    confirm: bool = True,
    tmp_path: Optional[Path] = None,
) -> dict:
    """
    执行完整接入流程：上传点表 → 解析 → 生成方案 → 确认 → 执行。

    参数:
        chat: ChatHelper 实例
        agent: AgentHandle 实例（config.json 产物从其 config_dir 读取）
        file_path: 点表文件路径
        upload_msg: 上传时的消息
        plan_msg: 方案生成消息
        confirm: 是否确认方案
        tmp_path: 历史参数，已废弃（保留兼容），产物路径统一取自 agent.config_dir

    返回:
        {
            "upload_text": str,
            "plan_text": str,
            "confirm_text": str or None,
            "interrupt_id": str or None,
            "config_json": dict or None,
        }
    """
    result: dict = {
        "upload_text": "",
        "plan_text": "",
        "confirm_text": None,
        "interrupt_id": None,
        "config_json": None,
    }

    # Step 1: 上传点表
    with chat.send_with_file(upload_msg, file_path) as stream:
        result["upload_text"] = stream.text_content()
    chat.record_response(result["upload_text"])

    # 直接从 CSV 文件解析设备数据（不依赖 SSE event 时序）
    result["upload_json"] = _parse_csv_to_device_json(file_path)

    # Step 2: 生成方案 — 嵌入上一步的设备信息
    upload_context = result["upload_text"]
    plan_with_context = f"{plan_msg}\n\n上一步解析结果:\n{upload_context}" if upload_context else plan_msg
    with chat.send(plan_with_context) as stream:
        result["plan_text"] = stream.text_content()
        result["interrupt_id"] = find_interrupt_id(stream)
    chat.record_response(result["plan_text"])

    # Step 3: 确认 — 嵌入设备 JSON
    if confirm:
        confirm_message = "[C4_BUTTON_CONFIRM] 确认"
        device_json = result.get("upload_json")
        if device_json:
            confirm_with_context = f"{confirm_message}\n\n{json.dumps(device_json)}"
        else:
            confirm_with_context = confirm_message
        with chat.send(confirm_with_context) as stream:
            result["confirm_text"] = stream.text_content()
        chat.record_response(result["confirm_text"])

    # Step 4: 读取 config.json 产物（与 conftest agent fixture 的 config_dir 一致）
    config_path = agent.config_dir / "config.json"
    if config_path.exists():
        try:
            result["config_json"] = json.loads(config_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass

    return result


def delete_device(chat: Any, agent: Any, device_name: str) -> None:
    """
    确定性删除设备：从 config.json 读取该设备实例的 instance.id，嵌入 changes JSON 到确认消息。

    避免依赖 LLM 将设备名映射为 instance.id（多轮长上下文中易出错）。
    """
    config_path = agent.config_dir / "config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))

    service_type: Optional[str] = None
    instance_id: Optional[str] = None
    for st, instances in config.items():
        if st == "c4_shm_manager" or not isinstance(instances, list):
            continue
        for inst in instances:
            name = str(inst.get("name", ""))
            if device_name in name or name in device_name:
                service_type = st
                instance_id = inst.get("id")
                break
        if instance_id:
            break
    assert instance_id, f"未在 config 中找到设备 {device_name} 的实例"

    with chat.send(f"停用 {device_name}") as s:
        s.text_content()

    changes = {
        "changes": [
            {
                "action": "delete",
                "service_type": service_type,
                "instance": {"id": instance_id},
            }
        ]
    }
    with chat.send(f"确认删除\n\n{json.dumps(changes, ensure_ascii=False)}") as s:
        s.text_content()


def modify_device(
    chat: Any,
    agent: Any,
    device_name: str,
    field: str,
    value: Any,
) -> None:
    """
    确定性修改设备：从 config.json 读取该设备实例的 instance.id，嵌入 modify changes JSON。

    避免依赖 LLM 将设备名映射为 instance.id（多轮长上下文中易出错，可能误生成 add）。
    """
    config_path = agent.config_dir / "config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))

    service_type: Optional[str] = None
    instance_id: Optional[str] = None
    for st, instances in config.items():
        if st == "c4_shm_manager" or not isinstance(instances, list):
            continue
        for inst in instances:
            name = str(inst.get("name", ""))
            if device_name in name or name in device_name:
                service_type = st
                instance_id = inst.get("id")
                break
        if instance_id:
            break
    assert instance_id, f"未在 config 中找到设备 {device_name} 的实例"

    with chat.send(f"修改 {device_name} 的 {field} 为 {value}") as s:
        s.text_content()

    changes = {
        "changes": [
            {
                "action": "modify",
                "service_type": service_type,
                "instance": {"id": instance_id, field: value},
            }
        ]
    }
    with chat.send(f"确认修改\n\n{json.dumps(changes, ensure_ascii=False)}") as s:
        s.text_content()
